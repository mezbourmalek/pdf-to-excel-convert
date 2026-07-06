from pathlib import Path
import re
import threading
import tkinter as tk
from tkinter import filedialog, messagebox

import customtkinter as ctk
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill


def clean_text(value):
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def normalize_text(value):
    value = clean_text(value).lower()
    value = value.replace("é", "e").replace("è", "e").replace("ê", "e")
    value = value.replace("à", "a").replace("ç", "c")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def row_to_text(row):
    return " ".join(clean_text(cell) for cell in row if clean_text(cell))


def is_empty_row(row):
    return all(clean_text(cell) == "" for cell in row)


def is_total_row(row):
    text = normalize_text(row_to_text(row))
    return (
        text.startswith("total")
        or " total " in f" {text} "
        or "totaux" in text
        or "sous-total" in text
        or "sous total" in text
    )


def find_header_index(table):
    keywords = [
        "t.qui", "n police", "n° police", "assure", "assuré",
        "effet", "expiration", "prime", "commission",
        "echeance", "emission", "dimension", "reliquat",
    ]

    best_index = None
    best_score = 0

    for index, row in enumerate(table):
        text = normalize_text(row_to_text(row))
        score = sum(1 for keyword in keywords if keyword in text)

        if score > best_score:
            best_score = score
            best_index = index

    if best_score >= 2:
        return best_index

    return None


def split_prefix_and_key(text_before_colon):
    text_before_colon = clean_text(text_before_colon)

    words = text_before_colon.split()
    if not words:
        return "", ""

    if len(words) == 1:
        return "", words[0]

    return " ".join(words[:-1]), words[-1]


def extract_metadata_pairs_from_text(text):
    pairs = []
    current_key = None
    current_value_parts = []

    text = clean_text(text)
    parts = [part for part in re.split(r"(\S+\s*:)", text) if clean_text(part)]

    for part in parts:
        part = clean_text(part)

        if part.endswith(":"):
            before_colon = part[:-1]
            prefix, new_key = split_prefix_and_key(before_colon)

            if prefix and current_key:
                current_value_parts.append(prefix)

            if current_key:
                value = " ".join(current_value_parts).strip()
                if value:
                    pairs.append((current_key, value))

            current_key = new_key
            current_value_parts = []
        else:
            if ":" in part:
                before_colon, after_colon = part.split(":", 1)
                prefix, new_key = split_prefix_and_key(before_colon)

                if prefix and current_key:
                    current_value_parts.append(prefix)

                if current_key:
                    value = " ".join(current_value_parts).strip()
                    if value:
                        pairs.append((current_key, value))

                current_key = new_key
                current_value_parts = []

                if after_colon.strip():
                    current_value_parts.append(clean_text(after_colon))
            elif current_key:
                current_value_parts.append(part)

    if current_key:
        value = " ".join(current_value_parts).strip()
        if value:
            pairs.append((current_key, value))

    return pairs


def extract_metadata_pairs_from_rows(metadata_rows):
    text = " ".join(row_to_text(row) for row in metadata_rows)
    return extract_metadata_pairs_from_text(text)


def get_metadata_pairs(page, table_object, table, header_index):
    metadata_rows = table[:header_index]
    pairs = extract_metadata_pairs_from_rows(metadata_rows)

    if pairs:
        return pairs

    table_top = table_object.bbox[1]
    y0 = max(0, table_top - 80)

    small_area_above_table = page.crop((0, y0, page.width, table_top))
    top_text = small_area_above_table.extract_text() or ""

    return extract_metadata_pairs_from_text(top_text)


def metadata_values(metadata_pairs):
    return [value for _, value in metadata_pairs]


def make_group_key(values):
    return tuple(normalize_text(value) for value in values)


def clean_row(row):
    return [clean_text(cell) for cell in row]


def sanitize_for_excel(value):
    """Empêche Excel d'interpréter une valeur comme une formule.

    Si la valeur commence par =, +, -, @ ou une tabulation, on la préfixe
    d'une apostrophe pour forcer une interprétation en texte brut
    (protection contre l'injection de formule / CSV injection).
    """
    if value and value[0] in ("=", "+", "-", "@", "\t"):
        return "'" + value
    return value


def write_table(sheet, start_row, table_rows):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="EDEDED")

    current_row = start_row

    for row_index, row in enumerate(table_rows):
        if is_empty_row(row):
            continue

        if is_total_row(row):
            continue

        cleaned = [sanitize_for_excel(cell_value) for cell_value in clean_row(row)]

        for column_index, value in enumerate(cleaned, start=1):
            cell = sheet.cell(row=current_row, column=column_index, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if row_index == 0:
                cell.font = Font(bold=True)
                cell.fill = header_fill

        current_row += 1

    return current_row


def autofit_columns(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            value = clean_text(cell.value)
            max_length = max(max_length, len(value))

        sheet.column_dimensions[column_letter].width = min(max_length + 2, 45)


def add_page_block(sheet, metadata_pairs, table_rows):
    start_row = sheet.max_row + 3 if sheet.max_row > 1 else 1
    current_row = start_row

    for pair_index in range(0, len(metadata_pairs), 2):
        left_key, left_value = metadata_pairs[pair_index]

        sheet.cell(row=current_row, column=1, value=sanitize_for_excel(f"{left_key} :"))
        sheet.cell(row=current_row, column=2, value=sanitize_for_excel(left_value))
        sheet.cell(row=current_row, column=1).font = Font(bold=True)

        if pair_index + 1 < len(metadata_pairs):
            right_key, right_value = metadata_pairs[pair_index + 1]

            sheet.cell(row=current_row, column=6, value=sanitize_for_excel(f"{right_key} :"))
            sheet.cell(row=current_row, column=7, value=sanitize_for_excel(right_value))
            sheet.cell(row=current_row, column=6).font = Font(bold=True)

        current_row += 1

    current_row += 1
    write_table(sheet, current_row, table_rows)
    autofit_columns(sheet)


def convert_pdf_to_excel(pdf_path, excel_path, progress_callback=None):
    workbook = Workbook()
    workbook.remove(workbook.active)

    groups = {}
    sheet_counter = 1

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)

        for page_number, page in enumerate(pdf.pages, start=1):
            found_tables = page.find_tables()

            for table_object in found_tables:
                table = table_object.extract()

                if not table:
                    continue

                header_index = find_header_index(table)

                if header_index is None:
                    continue

                metadata_pairs = get_metadata_pairs(
                    page,
                    table_object,
                    table,
                    header_index
                )

                values = metadata_values(metadata_pairs)
                group_key = make_group_key(values)

                if group_key not in groups:
                    sheet_name = f"Sheet {sheet_counter}"
                    groups[group_key] = workbook.create_sheet(sheet_name)
                    sheet_counter += 1

                table_rows = table[header_index:]
                add_page_block(groups[group_key], metadata_pairs, table_rows)

            if progress_callback:
                progress_callback(page_number, total_pages)

    if not workbook.sheetnames:
        raise ValueError("Aucun tableau utile n'a ete trouve dans ce PDF.")

    workbook.save(excel_path)


# ----------------------------------------------------------------------
# Interface graphique moderne (CustomTkinter)
# ----------------------------------------------------------------------

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")


class ConverterApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Convertisseur PDF → Excel")
        self.geometry("620x420")
        self.resizable(False, False)

        self.selected_pdf_path = None

        # ---- Conteneur principal ----
        self.container = ctk.CTkFrame(self, corner_radius=20)
        self.container.pack(fill="both", expand=True, padx=24, pady=24)

        # ---- Titre ----
        ctk.CTkLabel(
            self.container,
            text="Convertisseur PDF → Excel",
            font=ctk.CTkFont(size=22, weight="bold"),
        ).pack(pady=(28, 4))

        ctk.CTkLabel(
            self.container,
            text="Transforme tes relevés PDF en fichiers Excel propres et structurés",
            font=ctk.CTkFont(size=13),
            text_color="gray70",
        ).pack(pady=(0, 24))

        # ---- Zone de sélection de fichier (style "dropzone") ----
        self.drop_frame = ctk.CTkFrame(
            self.container,
            corner_radius=16,
            fg_color=("#f2f2f2", "#2a2d2e"),
            border_width=2,
            border_color=("#d0d0d0", "#3a3d3e"),
            height=110,
        )
        self.drop_frame.pack(fill="x", padx=30, pady=(0, 20))
        self.drop_frame.pack_propagate(False)

        self.file_label = ctk.CTkLabel(
            self.drop_frame,
            text="📄  Aucun fichier sélectionné",
            font=ctk.CTkFont(size=14),
            text_color="gray60",
        )
        self.file_label.pack(expand=True)

        self.choose_button = ctk.CTkButton(
            self.container,
            text="Choisir un fichier PDF",
            command=self.choose_pdf,
            height=42,
            corner_radius=10,
            font=ctk.CTkFont(size=14, weight="bold"),
        )
        self.choose_button.pack(fill="x", padx=30, pady=(0, 14))

        # ---- Bouton de conversion ----
        self.convert_button = ctk.CTkButton(
            self.container,
            text="Convertir et enregistrer en Excel",
            command=self.start_conversion,
            height=48,
            corner_radius=10,
            font=ctk.CTkFont(size=15, weight="bold"),
            fg_color="#2fa572",
            hover_color="#248a5c",
        )
        self.convert_button.pack(fill="x", padx=30, pady=(0, 18))

        # ---- Barre de progression + statut ----
        self.progress_bar = ctk.CTkProgressBar(self.container, height=10)
        self.progress_bar.set(0)
        self.progress_bar.pack(fill="x", padx=30, pady=(0, 8))

        self.status_label = ctk.CTkLabel(
            self.container,
            text="En attente d'un fichier…",
            font=ctk.CTkFont(size=12),
            text_color="gray60",
        )
        self.status_label.pack(pady=(0, 10))

    # ------------------------------------------------------------------
    def choose_pdf(self):
        file_path = filedialog.askopenfilename(
            title="Choisir un fichier PDF",
            filetypes=[("Fichiers PDF", "*.pdf")],
        )

        if file_path:
            self.selected_pdf_path = Path(file_path)
            self.file_label.configure(
                text=f"✅  {self.selected_pdf_path.name}",
                text_color=("black", "white"),
            )
            self.status_label.configure(text="Prêt à convertir.")
            self.progress_bar.set(0)

    def start_conversion(self):
        if not self.selected_pdf_path:
            messagebox.showwarning("Attention", "Choisis d'abord un fichier PDF.")
            return

        excel_file = filedialog.asksaveasfilename(
            title="Enregistrer le fichier Excel",
            defaultextension=".xlsx",
            filetypes=[("Fichiers Excel", "*.xlsx")],
            initialfile=f"{self.selected_pdf_path.stem}.xlsx",
        )

        if not excel_file:
            return

        self.convert_button.configure(state="disabled", text="Conversion en cours…")
        self.choose_button.configure(state="disabled")
        self.progress_bar.set(0)
        self.status_label.configure(text="Lecture du PDF…")

        thread = threading.Thread(
            target=self.run_conversion,
            args=(self.selected_pdf_path, Path(excel_file)),
            daemon=True,
        )
        thread.start()

    def run_conversion(self, pdf_path, excel_path):
        try:
            def on_progress(current, total):
                fraction = current / total if total else 0
                self.after(0, self.update_progress, fraction, current, total)

            convert_pdf_to_excel(pdf_path, excel_path, progress_callback=on_progress)
            self.after(0, self.conversion_done, True, None)
        except Exception as error:
            self.after(0, self.conversion_done, False, str(error))

    def update_progress(self, fraction, current, total):
        self.progress_bar.set(fraction)
        self.status_label.configure(text=f"Traitement de la page {current}/{total}…")

    def conversion_done(self, success, error_message):
        self.convert_button.configure(state="normal", text="Convertir et enregistrer en Excel")
        self.choose_button.configure(state="normal")

        if success:
            self.progress_bar.set(1)
            self.status_label.configure(text="Conversion terminée ✅")
            messagebox.showinfo("Succès", "Conversion terminée avec succès.")
        else:
            self.progress_bar.set(0)
            self.status_label.configure(text="Une erreur est survenue.")
            messagebox.showerror("Erreur", error_message)


if __name__ == "__main__":
    app = ConverterApp()
    app.mainloop()